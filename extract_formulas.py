import openpyxl
import os

# usage: python extract_formulas.py

def extract_formulas(filename):
    print(f"--- Extracting formulas from {filename} ---")
    try:
        # data_only=False ensures we get formulas, not evaluated values
        wb = openpyxl.load_workbook(filename, data_only=False)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        print(f"Cell {cell.coordinate}: {cell.value}")
                        
    except Exception as e:
        print(f"Error reading {filename}: {e}")

if __name__ == "__main__":
    # Check current directory and 'excel files' directory
    target_dir = os.path.join(os.getcwd(), 'excel files')
    if os.path.exists(target_dir):
        files = [os.path.join(target_dir, f) for f in os.listdir(target_dir) if f.endswith('.xlsx')]
    else:
        files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    if files:
        for file in files:
            print(f"I found {file}")
            extract_formulas(file)
    else:
        print("No .xlsx files found.")
